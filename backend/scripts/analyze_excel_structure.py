"""
Analiza la estructura exacta del Excel para entender cómo mapear los datos
"""
import sys
from pathlib import Path
import pandas as pd
import openpyxl

sys.path.insert(0, '/app')

excel_path = "/app/config/employee_master.xlsm"

print("=" * 60)
print("ANÁLISIS DETALLADO DEL ARCHIVO EXCEL")
print("=" * 60)

# Cargar con openpyxl para ver estructura más clara
wb = openpyxl.load_workbook(excel_path)
print(f"\n📄 Hojas disponibles ({len(wb.sheetnames)}):")
for i, sheet in enumerate(wb.sheetnames, 1):
    ws = wb[sheet]
    print(f"  {i}. {sheet} - {ws.max_row} filas, {ws.max_column} columnas")

# Analizar la hoja '派遣社員' en detalle
print(f"\n" + "=" * 60)
print(f"ANALIZANDO: 派遣社員 (Dispatch Employees)")
print("=" * 60)

ws = wb['派遣社員']

# Mostrar primeras 10 filas
print(f"\n📝 Primeras 10 filas (primeras 15 columnas):")
for row_idx in range(1, 11):
    values = []
    for col_idx in range(1, 16):
        cell = ws.cell(row_idx, col_idx)
        val = cell.value
        if val is None:
            val = "---"
        values.append(str(val)[:15])  # Truncar a 15 caracteres
    print(f"  Fila {row_idx:2d}: {' | '.join(values)}")

# Buscar dónde empiezan los datos reales
print(f"\n🔍 Buscando inicio de datos útiles...")
for row_idx in range(1, min(20, ws.max_row + 1)):
    row_data = [ws.cell(row_idx, col).value for col in range(1, 6)]
    print(f"  Fila {row_idx}: {row_data}")

print("\n✅ Análisis completado. Revisa arriba para entender la estructura.")
