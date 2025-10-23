#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Analizar employee_master.xlsm para generar estructuras de BD
"""
import openpyxl
import sys
import json

def analyze_sheet(wb, sheet_name, header_row=2):
    """Analizar una hoja del Excel"""
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]

    # Obtener encabezados
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value and str(cell_value).strip():
            headers.append({
                'col_num': col,
                'name': str(cell_value).strip()
            })

    # Obtener muestras de datos
    samples = []
    for row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        row_data = {}
        for h in headers:
            val = ws.cell(row, h['col_num']).value
            if val is not None:
                row_data[h['name']] = str(val)[:50]
        if row_data:
            samples.append(row_data)

    return {
        'sheet_name': sheet_name,
        'total_rows': ws.max_row,
        'total_cols': ws.max_column,
        'headers': headers,
        'sample_data': samples[:3]
    }

def main():
    sys.stdout.reconfigure(encoding='utf-8')

    wb = openpyxl.load_workbook(
        'frontend-nextjs/app/factories/employee_master.xlsm',
        read_only=True,
        data_only=True
    )

    sheets_to_analyze = [
        ('派遣社員', 2),
        ('請負社員', 2),
        ('スタッフ', 2)
    ]

    results = {}

    for sheet_name, header_row in sheets_to_analyze:
        print(f'\n{"="*60}')
        print(f'ANALIZANDO: {sheet_name}')
        print(f'{"="*60}\n')

        data = analyze_sheet(wb, sheet_name, header_row)
        if data:
            results[sheet_name] = data

            print(f'Filas totales: {data["total_rows"]}')
            print(f'Columnas totales: {data["total_cols"]}')
            print(f'Encabezados encontrados: {len(data["headers"])}\n')

            print('COLUMNAS:')
            for h in data['headers']:
                print(f'  {h["col_num"]:2d}. {h["name"]}')

            if data['sample_data']:
                print(f'\nMUESTRA DE DATOS (primeras {len(data["sample_data"])} filas):')
                for i, row in enumerate(data['sample_data'], 1):
                    print(f'\n  Fila {i}:')
                    for key, val in list(row.items())[:10]:
                        print(f'    {key}: {val}')

    # Guardar resultados en JSON
    with open('excel_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f'\n\n{"="*60}')
    print('ANÁLISIS COMPLETO - Guardado en excel_analysis.json')
    print(f'{"="*60}')

if __name__ == '__main__':
    main()
